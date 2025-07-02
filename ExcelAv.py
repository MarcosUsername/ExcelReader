from openpyxl import load_workbook
import pandas as pd

# State sheet name
sheet_name = 'test raw data'

# Load workbook and stated sheet
file_path = 'D:/Documents/Work/Monica Data Evaluation/2025-5-3 L-03 71Cycles.csv.xlsx'  # Replace file path
wb = load_workbook(filename=file_path, data_only=True)
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found.")
sheet = wb[sheet_name]

# List cell ranges
ranges = {
    'Range1': 'D2:D3000',
    'Range2': 'I2:I3000',
    'Range3': 'N2:N3000',
    'Range4': 'S2:S3000',
    'Range5': 'W2:W3000',
    'Range6': 'AC2:AC3000',
}

# Collect results
results = []

for name, cell_range in ranges.items():
    cells = sheet[cell_range]
    values = [cell[0].value for cell in cells if cell[0].value not in (0, None)]

    total = sum(values)
    count = len(values)
    average = total / count if count > 0 else 0

    results.append({
        #'Name': name,
        'Range': cell_range,
        'Total': total,
        'Av': average
    })

# Convert results
output_df = pd.DataFrame(results)

# Save
output_path = 'D:/Documents/output.xlsx'  # Save location
output_df.to_excel(output_path, index=False)

print(f"Summary saved to '{output_path}'")


#=================================================================================================

# Author: Marco Silvestri
# Email: Marco.Silvestri@CabinAir.com
# Date Created: 02-07-25
# Last Modified: 02-07-25
# Description: This script is to take data from an Excel file, calculate the total and average for specified ranges, and save the results to a new Excel file.
# Licence: MIT License

#  __  __  ____  
# |  \/  |/ ___| 
# | |\/| |\___ \ 
# | |  | | ___) |
# |_|  |_||____/ 